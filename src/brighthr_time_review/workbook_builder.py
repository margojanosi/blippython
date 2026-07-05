"""Excel workbook builder for BrightHR Time Review MVP.

Generates a formatted multi-tab Excel workbook from the normalised data
and detected exceptions.

Tabs produced:
  1. Instructions
  2. Exception Report   (main review tab with conditional formatting)
  3. Summary            (dashboard counts)
  4. Rules Used         (active thresholds from config)
  5. Raw Normalized Data
  6. Review Log
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from brighthr_time_review.exceptions import ExceptionRecord

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour palette (openpyxl uses ARGB hex, first two = alpha=FF)
# ---------------------------------------------------------------------------
COLOUR_RED_FILL = PatternFill("solid", fgColor="FFFF9999")      # High severity
COLOUR_AMBER_FILL = PatternFill("solid", fgColor="FFFFD699")    # Medium severity
COLOUR_GREEN_FILL = PatternFill("solid", fgColor="FFCCFFCC")    # Low severity
COLOUR_HEADER_FILL = PatternFill("solid", fgColor="FF1F4E79")   # Dark blue header
COLOUR_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
COLOUR_TITLE_FONT = Font(size=14, bold=True, color="FF1F4E79")
COLOUR_WARN_FILL = PatternFill("solid", fgColor="FFFFF2CC")     # Warning yellow

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _set_col_widths(ws: Worksheet, min_width: int = 12, max_width: int = 45) -> None:
    """Auto-size columns based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, val_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def _write_header_row(ws: Worksheet, headers: list[str], row: int = 1) -> None:
    """Write a styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = COLOUR_HEADER_FONT
        cell.fill = COLOUR_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Tab 1 – Instructions
# ---------------------------------------------------------------------------

def _build_instructions_tab(ws: Worksheet) -> None:
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 80

    title_cell = ws["B2"]
    title_cell.value = "BrightHR Time Review – Exception Workbook"
    title_cell.font = Font(size=16, bold=True, color="FF1F4E79")

    rows = [
        (4, "HOW TO USE THIS WORKBOOK", Font(bold=True, size=12)),
        (6, "Step 1:  Export the BrightHR/Blip timesheet CSV from the BrightHR portal.", None),
        (7, "Step 2:  Save the CSV file into the  data/input/  folder.", None),
        (8, "Step 3:  Run the review tool with:  python -m brighthr_time_review.main", None),
        (9, "Step 4:  Open this workbook and review the  Exception Report  tab.", None),
        (10, "Step 5:  For each flagged row, confirm or correct the record manually in BrightHR.", None),
        (11, "Step 6:  Update the  Status  and  Reviewer Notes  columns, then save this workbook as your payroll audit trail.", None),
        (13, "⚠️  IMPORTANT – WORKBOOK BOUNDARIES", Font(bold=True, size=12, color="FFCC0000")),
        (14, "• This workbook does NOT approve payroll.", None),
        (15, "• This workbook does NOT correct records in BrightHR.", None),
        (16, "• This workbook does NOT submit anything to Dayforce.", None),
        (17, "• Every flagged exception must be reviewed and confirmed by a human reviewer.", None),
        (18, "• BrightHR is the authoritative source of truth for all time records.", None),
        (20, "TABS IN THIS WORKBOOK", Font(bold=True, size=12)),
        (21, "Exception Report   – Main review tab. Work through every row.", None),
        (22, "Summary            – High-level counts and breakdown by severity / type.", None),
        (23, "Rules Used         – Active detection thresholds loaded from config.", None),
        (24, "Raw Normalized Data – Cleaned version of the BrightHR export.", None),
        (25, "Review Log         – Run metadata and processing notes.", None),
    ]

    for row_num, text, font_override in rows:
        cell = ws.cell(row=row_num, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True)
        if font_override:
            cell.font = font_override

    ws.row_dimensions[2].height = 25


# ---------------------------------------------------------------------------
# Tab 2 – Exception Report
# ---------------------------------------------------------------------------

EXCEPTION_REPORT_COLS = [
    "Employee Name",
    "Employee ID",
    "Work Date",
    "Exception Type",
    "Severity",
    "Clock In",
    "Clock Out",
    "Total Shift Hours",
    "Break Minutes",
    "Raw Value",
    "Rule Triggered",
    "Suggested Action",
    "Source Row Number",
    "Status",
    "Reviewer Notes",
]


def _build_exception_report_tab(
    ws: Worksheet,
    exceptions: list[ExceptionRecord],
) -> None:
    ws.title = "Exception Report"

    _write_header_row(ws, EXCEPTION_REPORT_COLS)
    ws.freeze_panes = "A2"

    severity_col = EXCEPTION_REPORT_COLS.index("Severity") + 1  # 1-based

    for row_idx, exc in enumerate(exceptions, start=2):
        row_data = exc.as_dict()
        for col_idx, col_name in enumerate(EXCEPTION_REPORT_COLS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Apply table formatting
    if exceptions:
        last_row = len(exceptions) + 1
        last_col = get_column_letter(len(EXCEPTION_REPORT_COLS))
        table_ref = f"A1:{last_col}{last_row}"
        table = Table(displayName="ExceptionReport", ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Conditional formatting for Severity column
        sev_col_letter = get_column_letter(severity_col)
        data_range = f"{sev_col_letter}2:{sev_col_letter}{last_row}"

        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="equal", formula=['"High"'], fill=COLOUR_RED_FILL),
        )
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="equal", formula=['"Medium"'], fill=COLOUR_AMBER_FILL),
        )
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator="equal", formula=['"Low"'], fill=COLOUR_GREEN_FILL),
        )

    ws.auto_filter.ref = ws.dimensions
    _set_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 3 – Summary
# ---------------------------------------------------------------------------

def _build_summary_tab(
    ws: Worksheet,
    df_normalized: pd.DataFrame,
    exceptions: list[ExceptionRecord],
) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    title_cell = ws["A1"]
    title_cell.value = "BrightHR Time Review – Summary"
    title_cell.font = COLOUR_TITLE_FONT

    total_raw = len(df_normalized)
    flagged_rows = df_normalized[
        df_normalized["normalization_warnings"].str.len() > 0
    ]
    total_exceptions = len(exceptions)
    open_exc = sum(1 for e in exceptions if e.status == "Open")
    closed_exc = sum(1 for e in exceptions if e.status == "Closed")
    clean_records = total_raw - len(
        set(e.source_row for e in exceptions)
    )

    def _write_kv(row: int, label: str, value: Any, bold: bool = False) -> None:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.alignment = Alignment(horizontal="left")
        vc.alignment = Alignment(horizontal="right")
        if bold:
            lc.font = Font(bold=True)
            vc.font = Font(bold=True)

    row = 3
    _write_kv(row, "Total Raw Records Reviewed", total_raw, bold=True)
    row += 1
    _write_kv(row, "Total Clean Records", clean_records)
    row += 1
    _write_kv(row, "Total Records with Normalization Warnings", len(flagged_rows))
    row += 1
    _write_kv(row, "Total Exceptions Detected", total_exceptions, bold=True)
    row += 1
    _write_kv(row, "  Open Exceptions", open_exc)
    row += 1
    _write_kv(row, "  Closed Exceptions", closed_exc)

    row += 2
    ws.cell(row=row, column=1, value="Exceptions by Severity").font = Font(bold=True)
    row += 1
    by_severity: dict[str, int] = {}
    for exc in exceptions:
        by_severity[exc.severity] = by_severity.get(exc.severity, 0) + 1
    for sev in ["High", "Medium", "Low"]:
        _write_kv(row, f"  {sev}", by_severity.get(sev, 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Exceptions by Type").font = Font(bold=True)
    row += 1
    by_type: dict[str, int] = {}
    for exc in exceptions:
        by_type[exc.exception_type] = by_type.get(exc.exception_type, 0) + 1
    for exc_type, count in sorted(by_type.items()):
        _write_kv(row, f"  {exc_type}", count)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Exceptions by Employee").font = Font(bold=True)
    row += 1
    by_emp: dict[str, int] = {}
    for exc in exceptions:
        by_emp[exc.employee_name] = by_emp.get(exc.employee_name, 0) + 1
    for emp, count in sorted(by_emp.items()):
        _write_kv(row, f"  {emp}", count)
        row += 1


# ---------------------------------------------------------------------------
# Tab 4 – Rules Used
# ---------------------------------------------------------------------------

RULES_COLS = ["Rule Key", "Rule Name", "Value", "Unit", "Severity", "Enabled", "Description"]


def _build_rules_tab(ws: Worksheet, rules: dict) -> None:
    ws.title = "Rules Used"
    _write_header_row(ws, RULES_COLS)
    ws.freeze_panes = "A2"

    row_idx = 2
    for key, value in rules.items():
        if isinstance(value, dict):
            ws.cell(row=row_idx, column=1, value=key)
            ws.cell(row=row_idx, column=2, value=value.get("description", ""))
            ws.cell(row=row_idx, column=3, value=value.get("value", ""))
            ws.cell(row=row_idx, column=4, value=value.get("unit", ""))
            ws.cell(row=row_idx, column=5, value=value.get("severity", ""))
            ws.cell(row=row_idx, column=6, value=str(value.get("enabled", True)))
            ws.cell(row=row_idx, column=7, value=value.get("description", ""))
            row_idx += 1

    _set_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 5 – Raw Normalized Data
# ---------------------------------------------------------------------------

RAW_DATA_COLS = [
    "Employee Name",
    "Employee ID",
    "Work Date",
    "Day of Week",
    "Is Weekend",
    "Clock In",
    "Clock Out",
    "Break Start",
    "Break End",
    "Break Minutes",
    "Calculated Shift Hours",
    "Calculated Paid Hours",
    "Has Clock In",
    "Has Clock Out",
    "Has Break",
    "Source Row Number",
    "Normalization Warning",
]


def _build_raw_data_tab(ws: Worksheet, df: pd.DataFrame) -> None:
    ws.title = "Raw Normalized Data"
    _write_header_row(ws, RAW_DATA_COLS)
    ws.freeze_panes = "A2"

    col_map = {
        "Employee Name": "employee_name",
        "Employee ID": "employee_id",
        "Work Date": "work_date",
        "Day of Week": "day_of_week",
        "Is Weekend": "is_weekend",
        "Clock In": "clock_in",
        "Clock Out": "clock_out",
        "Break Start": "break_start",
        "Break End": "break_end",
        "Break Minutes": "calc_break_minutes",
        "Calculated Shift Hours": "calc_shift_hours",
        "Calculated Paid Hours": "calc_paid_hours",
        "Has Clock In": "has_clock_in",
        "Has Clock Out": "has_clock_out",
        "Has Break": "has_break",
        "Source Row Number": "source_row",
        "Normalization Warning": "normalization_warnings",
    }

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, display_col in enumerate(RAW_DATA_COLS, start=1):
            src_col = col_map.get(display_col, display_col)
            val = row.get(src_col, "")
            # Convert Timestamps to strings for Excel
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%H:%M")
            elif pd.isna(val):
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val).alignment = Alignment(
                vertical="top"
            )

    _set_col_widths(ws)


# ---------------------------------------------------------------------------
# Tab 6 – Review Log
# ---------------------------------------------------------------------------

def _build_review_log_tab(
    ws: Worksheet,
    input_path: Path,
    output_path: Path,
    df_normalized: pd.DataFrame,
    exceptions: list[ExceptionRecord],
    rules: dict,
    warnings: list[str] | None = None,
) -> None:
    ws.title = "Review Log"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 55

    title_cell = ws["A1"]
    title_cell.value = "BrightHR Time Review – Run Log"
    title_cell.font = COLOUR_TITLE_FONT

    def _kv(row: int, label: str, value: Any) -> None:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row, column=2, value=value)
        cell.alignment = Alignment(wrap_text=True)

    row = 3
    _kv(row, "Run Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row += 1
    _kv(row, "Input File", str(input_path))
    row += 1
    _kv(row, "Output File", str(output_path))
    row += 1
    _kv(row, "Source Data Row Count", len(df_normalized))
    row += 1
    _kv(row, "Normalised Record Count", len(df_normalized))
    row += 1
    _kv(row, "Exception Count", len(exceptions))
    row += 1

    enabled_rules = [k for k, v in rules.items() if isinstance(v, dict) and v.get("enabled")]
    _kv(row, "Enabled Rules", ", ".join(enabled_rules))
    row += 2

    if warnings:
        ws.cell(row=row, column=1, value="Processing Warnings").font = Font(
            bold=True, color="FFCC0000"
        )
        row += 1
        for w in warnings:
            ws.cell(row=row, column=1, value=w).fill = COLOUR_WARN_FILL
            row += 1
        row += 1

    row += 1
    reminder = ws.cell(
        row=row,
        column=1,
        value=(
            "⚠️  Reminder: All corrections must be made manually in BrightHR.  "
            "This workbook does not modify any payroll or BrightHR records."
        ),
    )
    reminder.font = Font(bold=True, color="FFCC0000")
    reminder.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)


# ---------------------------------------------------------------------------
# Master builder
# ---------------------------------------------------------------------------

def build_workbook(
    df_normalized: pd.DataFrame,
    exceptions: list[ExceptionRecord],
    rules: dict,
    input_path: Path,
    output_path: Path,
    warnings: list[str] | None = None,
) -> Path:
    """Build and save the multi-tab Excel workbook.

    Args:
        df_normalized: Normalised DataFrame from the normalizer.
        exceptions:    List of detected ExceptionRecord objects.
        rules:         Parsed exception rules dictionary.
        input_path:    Original CSV input path (for logging).
        output_path:   Destination .xlsx path.
        warnings:      Optional list of processing warnings to include.

    Returns:
        Resolved path of the saved workbook.
    """
    logger.info("Building workbook with %d exceptions …", len(exceptions))

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Build each tab
    _build_instructions_tab(wb.create_sheet("Instructions"))
    _build_exception_report_tab(wb.create_sheet("Exception Report"), exceptions)
    _build_summary_tab(wb.create_sheet("Summary"), df_normalized, exceptions)
    _build_rules_tab(wb.create_sheet("Rules Used"), rules)
    _build_raw_data_tab(wb.create_sheet("Raw Normalized Data"), df_normalized)
    _build_review_log_tab(
        wb.create_sheet("Review Log"),
        input_path=input_path,
        output_path=output_path,
        df_normalized=df_normalized,
        exceptions=exceptions,
        rules=rules,
        warnings=warnings or [],
    )

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    logger.info("Workbook saved to %s", out)
    return out.resolve()
