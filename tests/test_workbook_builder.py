"""Tests for the workbook builder module."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from brighthr_time_review.config import load_exception_rules
from brighthr_time_review.loader import load_csv
from brighthr_time_review.normalizer import normalize
from brighthr_time_review.rules import detect_all
from brighthr_time_review.workbook_builder import build_workbook

_RULES_PATH = (
    Path(__file__).resolve().parents[1] / "config" / "exception_rules.yml"
)
_SAMPLE_CSV = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "input"
    / "sample_bright_hr_export.csv"
)

EXPECTED_TABS = [
    "Instructions",
    "Exception Report",
    "Summary",
    "Rules Used",
    "Raw Normalized Data",
    "Review Log",
]


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    """Generate the workbook once and return the path for all tests."""
    rules = load_exception_rules(_RULES_PATH)
    df = normalize(load_csv(_SAMPLE_CSV))
    exceptions = detect_all(df, rules, [])
    out = tmp_path / "test_output.xlsx"
    build_workbook(
        df_normalized=df,
        exceptions=exceptions,
        rules=rules,
        input_path=_SAMPLE_CSV,
        output_path=out,
    )
    return out


def test_workbook_file_created(workbook_path: Path) -> None:
    assert workbook_path.exists()
    assert workbook_path.stat().st_size > 0


def test_workbook_is_valid_xlsx(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    assert wb is not None


def test_workbook_contains_expected_tabs(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    for expected_tab in EXPECTED_TABS:
        assert expected_tab in wb.sheetnames, f"Missing tab: {expected_tab}"


def test_exception_report_has_header_row(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Exception Report"]
    headers = [cell.value for cell in ws[1]]
    assert "Employee Name" in headers
    assert "Exception Type" in headers
    assert "Severity" in headers
    assert "Status" in headers


def test_exception_report_has_data_rows(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Exception Report"]
    # Row 1 is header; there should be at least one data row
    data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]
    assert len(data_rows) > 0


def test_summary_tab_exists_and_has_content(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Summary"]
    non_empty = [row for row in ws.iter_rows(values_only=True) if any(row)]
    assert len(non_empty) > 0


def test_rules_used_tab_has_content(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Rules Used"]
    # Should have at least a header row and one data row
    rows_with_data = [row for row in ws.iter_rows(values_only=True) if any(row)]
    assert len(rows_with_data) >= 2


def test_raw_normalized_data_tab_has_content(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Raw Normalized Data"]
    rows_with_data = [row for row in ws.iter_rows(values_only=True) if any(row)]
    assert len(rows_with_data) >= 2  # header + at least one data row


def test_review_log_tab_has_run_timestamp(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Review Log"]
    all_cell_values = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                all_cell_values.append(str(cell.value))
    assert any("Run Timestamp" in v for v in all_cell_values)


def test_exceptions_appear_on_report(workbook_path: Path) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Exception Report"]
    exception_types = [
        ws.cell(row=r, column=4).value
        for r in range(2, ws.max_row + 1)
    ]
    exception_types = [t for t in exception_types if t]
    assert len(exception_types) > 0


def test_workbook_generation_with_empty_exceptions(tmp_path: Path) -> None:
    """Workbook should be generated even when no exceptions are found."""
    import pandas as pd
    from brighthr_time_review.workbook_builder import build_workbook as build_wb

    rules = load_exception_rules(_RULES_PATH)
    df = normalize(load_csv(_SAMPLE_CSV))
    out = tmp_path / "empty_exc.xlsx"
    build_wb(df_normalized=df, exceptions=[], rules=rules,
             input_path=_SAMPLE_CSV, output_path=out)
    assert out.exists()
    wb = load_workbook(out)
    assert "Exception Report" in wb.sheetnames
